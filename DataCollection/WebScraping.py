
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from os import path
from datetime import date
from datetime import datetime
import re

#If a workbook with data already exist, open and save data to it, else create a new workbook to work from
file_name = 'WeatherData1.xlsx'

if path.exists(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Location"
    sheet["B1"] = "Date"
    sheet["C1"] = "Time"
    sheet["D1"] = "Weather"
    sheet["E1"] = "Temperature (°C)"
    sheet["F1"] = "Humidity (%)"
    sheet["G1"] = "Wind (km/h)"
    sheet["H1"] = "Pressure (mb)"
    sheet["I1"] = "PM2.5 (µg/m³)"
    sheet["J1"] = "Pm10 (µg/m³)"
    sheet["K1"] = "AQI Value"

#Adding date, time and location to sheet
today = date.today()
date = today.strftime("%d/%m/%Y")
sheet.cell(column=2, row=sheet.max_row+1, value=date)
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
sheet.cell(column=3, row=sheet.max_row, value=current_time)
location = "stampfenbachstrasse" #Change if the location is changed
sheet.cell(column=1, row=sheet.max_row, value=location)

#Retrieve information from IQAIR
url = ('https://www.iqair.com/us/switzerland/zurich/zurich-stampfenbachstrasse')
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.set_window_size(480, 320)
# maximize window
driver.maximize_window()
driver.get(url)

#Give browser time to load (increase time value for less stable internet conncetions)
wait = WebDriverWait(driver, 10)

#Retrieve source information of the page
source = driver.page_source

# quit the driver
driver.quit()

#Creating usable HTML from the source
soup = BeautifulSoup(source, 'lxml')

#Find information regarding the general weather - search for table named "What is the current weather near Zürich Stampfenbachstrasse, Zurich?"
#This will need to be changed if the location is changed
table_general_weather = soup.find('table', {'title': "What is the current weather near Zürich Stampfenbachstrasse, Zurich?"})
given_data = ['Weather', 'Temperature', 'Humidity', 'Wind', 'Pressure']
info = table_general_weather.get_text()

#extracting numerical values from the information from the table
weather = re.search(given_data[0]+'(.*)'+given_data[1], info).group(1)
temp = re.search(given_data[1]+'(.*)'+given_data[2], info).group(1)
humidity = re.search(given_data[2]+'(.*)'+given_data[3], info).group(1)
wind = re.search(given_data[3]+'(.*)'+given_data[4], info).group(1)
pressure = re.search(given_data[4]+'(.*)', info).group(1)

sheet.cell(column=4, row=sheet.max_row, value=weather)
sheet.cell(column=5, row=sheet.max_row, value=temp.strip("°C"))
sheet.cell(column=6, row=sheet.max_row, value=humidity.strip("%"))
sheet.cell(column=7, row=sheet.max_row, value=wind.strip(" km/h"))
sheet.cell(column=8, row=sheet.max_row, value=pressure.strip(" mb"))

#Find information regarding the pollutants - search using the class type as there are two tables with the same title from which need to be scraped
#This will need to be changed if the location is changed
table_detailed_quality = soup.find('table', {'class': "aqi-overview-detail__other-pollution-table"})
given_data1 = ['PM2.5', 'pm10', 'o3', 'NO2']
pollutants_detail = table_detailed_quality.findAll('tr')
info = ''
for i in pollutants_detail:
    info += str(i.get_text())

#Not all the data is present so we have to see if the data can be collected:
if info.find('PM2.5') != -1:
    pm2_5 = re.search('PM2.5 (.*) ', info).group(1)
    pm2_5Value = pm2_5.split(' ')[1]
    sheet.cell(column=9, row=sheet.max_row, value=pm2_5Value)
else:
    sheet.cell(column=9, row=sheet.max_row, value='N/A')

if info.find('pm10') != -1:
    pm10 = re.search('pm10 (.*) ', info).group(1)
    pm10Value = pm10.split(' ')[1]
    sheet.cell(column=10, row=sheet.max_row, value=pm10Value)
else:
    sheet.cell(column=10, row=sheet.max_row, value='N/A')

try:
    aqi_value = soup.find('p', {'class': "aqi-value__value"}).get_text()
    sheet.cell(column=11, row=sheet.max_row, value=aqi_value)
except:
    sheet.cell(column=11, row=sheet.max_row, value="N/A")

#Save .xlxs workbook
workbook.save(filename=file_name)

